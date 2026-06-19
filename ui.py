import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import threading
import time
import datetime
from PIL import Image, ImageTk
import csv
import openpyxl
from openpyxl.styles import Font
import automate2  # Imports your existing automation script
from fundrel_automation.core.accounts import load_accounts as load_accounts_file
from fundrel_automation.core.accounts import prepare_account_config
from fundrel_automation.core.accounts import save_accounts as save_accounts_file
from fundrel_automation.core.config import load_settings, parse_proxies, parse_referrals, save_settings
from fundrel_automation.core.logging_config import get_logger
from fundrel_automation.core.paths import ACCOUNTS_PATH, workspace_path
from fundrel_automation.core.results import (
    apply_outcome_to_account,
    outcome_from_exception,
    outcome_from_result,
)
from fundrel_automation.core.status import TAB_STATUS_MAP, classify_account, reset_to_pending

logger = get_logger(__name__)

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tw = None
        self.id = None
        self.widget.bind("<Enter>", self.enter, add="+")
        self.widget.bind("<Leave>", self.leave, add="+")
        self.widget.bind("<ButtonPress>", self.leave, add="+")

    def enter(self, event=None):
        self.unschedule()
        self.id = self.widget.after(500, self.show)

    def leave(self, event=None):
        self.unschedule()
        self.hide()

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_:
            self.widget.after_cancel(id_)

    def show(self):
        if not self.text:
            return
        self.hide()
        x = self.widget.winfo_rootx() + 15
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tw = tk.Toplevel(self.widget)
        self.tw.wm_overrideredirect(True)
        self.tw.attributes("-topmost", True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tw, text=self.text, justify='left',
                         background='#ffffe0', relief='solid', borderwidth=1,
                         font=("Segoe UI", 9, "normal"), padx=5, pady=3)
        label.pack(ipadx=1)

    def hide(self):
        if self.tw:
            self.tw.destroy()
            self.tw = None

def add_tooltip(widget, text):
    ToolTip(widget, text)

ACCOUNTS_FILE = ACCOUNTS_PATH

class AutoUI:
    def __init__(self, root):
        self.root = root
        self.root.title("🚀 F Automation Control Panel")
        self.root.geometry("1200x700")
        self.root.configure(bg="#f4f6f9")
        
        self.accounts = []
        self.stop_requested = False
        self.is_running = False
        self.thread = None
        
        # Modern UI styling
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background="#f4f6f9", borderwidth=0)
        style.configure("TNotebook.Tab", font=("Segoe UI", 11, "bold"), padding=[15, 5], background="#e1e4e8", foreground="#333")
        style.map("TNotebook.Tab", background=[("selected", "#ffffff")], foreground=[("selected", "#0056b3")])
        
        style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"), background="#0056b3", foreground="white", borderwidth=0)
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30, background="#ffffff", fieldbackground="#ffffff", borderwidth=0)
        style.map('Treeview', background=[('selected', '#cce5ff')], foreground=[('selected', '#000000')])
        
        self.load_accounts()
        self.build_ui()
        self.refresh_lists()

    def load_accounts(self):
        try:
            self.accounts = load_accounts_file(ACCOUNTS_FILE)
        except FileNotFoundError:
            self.accounts = []
        except Exception as e:
            self.accounts = []
            logger.exception("Failed to load accounts")
            messagebox.showerror("Action needed", f"Failed to load accounts: {e}")

    def save_accounts(self):
        save_accounts_file(self.accounts, ACCOUNTS_FILE)

    def build_ui(self):
        # --- HEADER ---
        header = tk.Frame(self.root, bg="#0056b3", pady=15)
        header.pack(side=tk.TOP, fill=tk.X)
        
        title_lbl = tk.Label(header, text="FanDuel AutoBot Dashboard", font=("Segoe UI", 16, "bold"), bg="#0056b3", fg="white")
        title_lbl.pack(side=tk.LEFT, padx=20)

        # --- TOOLBAR ---
        toolbar = tk.Frame(self.root, bg="#ffffff", pady=10, padx=10, relief=tk.FLAT)
        toolbar.pack(side=tk.TOP, fill=tk.X, pady=(0, 10))
        
        btn_font = ("Segoe UI", 10, "bold")
        
        def create_btn(parent, text, cmd, bg_col, fg_col="black", tooltip_text=""):
            btn = tk.Button(parent, text=text, command=cmd, bg=bg_col, fg=fg_col, font=btn_font, 
                            relief=tk.FLAT, padx=15, pady=5, cursor="hand2")
            # Hover effects
            btn.bind("<Enter>", lambda e: btn.config(bg="#d1d5db" if bg_col == "#e2e8f0" else bg_col), add="+")
            btn.bind("<Leave>", lambda e: btn.config(bg=bg_col), add="+")
            if tooltip_text:
                add_tooltip(btn, tooltip_text)
            return btn

        create_btn(toolbar, "Upload JSON", self.upload_json, "#e2e8f0", tooltip_text="Load accounts from a .json file").pack(side=tk.LEFT, padx=5)
        create_btn(toolbar, "Upload Excel", self.upload_excel, "#e2e8f0", tooltip_text="Load accounts from a .xlsx or .xls file").pack(side=tk.LEFT, padx=5)
        create_btn(toolbar, "Add Account", self.add_account, "#e2e8f0", tooltip_text="Manually enter a single account's details").pack(side=tk.LEFT, padx=5)
        create_btn(toolbar, "Settings", self.open_settings, "#ffc107", tooltip_text="Configure proxies, referrals, and Edge browser path").pack(side=tk.LEFT, padx=5)
        
        separator = tk.Frame(toolbar, width=2, bg="#cbd5e1")
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=15, pady=2)
        
        create_btn(toolbar, "Run Pending", self.run_pending, "#28a745", "white", tooltip_text="Start automating all accounts in the Pending tab").pack(side=tk.LEFT, padx=5)
        create_btn(toolbar, "Run Selected", self.run_selected, "#17a2b8", "white", tooltip_text="Start automating ONLY the accounts you have highlighted in the Pending tab").pack(side=tk.LEFT, padx=5)
        
        self.stop_btn = create_btn(toolbar, "STOP EVERYTHING", self.stop_automation, "#dc3545", "white", tooltip_text="EMERGENCY BRAKE: Stop the bot immediately and close the browser")
        self.stop_btn.config(state=tk.DISABLED)
        self.stop_btn.pack(side=tk.RIGHT, padx=10)
        
        # Status panel
        status_frame = tk.Frame(toolbar, bg="#ffffff")
        status_frame.pack(side=tk.RIGHT, padx=20)
        
        self.status_lbl = tk.Label(status_frame, text="Idle", fg="#28a745", font=("Segoe UI", 11, "bold"), bg="#ffffff")
        self.status_lbl.pack(side=tk.TOP, anchor=tk.E)
        
        self.warning_lbl = tk.Label(status_frame, text="", fg="#dc3545", font=("Segoe UI", 9, "bold"), bg="#ffffff")
        self.warning_lbl.pack(side=tk.BOTTOM, anchor=tk.E)

        # --- TABS ---
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        self.tab_pending = ttk.Frame(self.notebook)
        self.tab_created = ttk.Frame(self.notebook)
        self.tab_failed = ttk.Frame(self.notebook)
        self.tab_skipped = ttk.Frame(self.notebook)
        self.tab_another_account = ttk.Frame(self.notebook)
        self.tab_service_unavailable = ttk.Frame(self.notebook)
        self.tab_unable_to_verify = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_pending, text="  Pending  ")
        self.notebook.add(self.tab_created, text="  Created  ")
        self.notebook.add(self.tab_failed, text="  Failed  ")
        self.notebook.add(self.tab_skipped, text="  Skipped  ")
        self.notebook.add(self.tab_another_account, text="  Another Account  ")
        self.notebook.add(self.tab_service_unavailable, text="  Service Unavailable  ")
        self.notebook.add(self.tab_unable_to_verify, text="  Unable to Verify  ")
        
        # Tooltips for notebook tabs
        tab_tooltips = {
            0: "Accounts that haven't been run yet, or were sent back to try again.",
            1: "Accounts that successfully reached the final success screen.",
            2: "Accounts that encountered an error while the bot was running.",
            3: "Accounts the bot skipped (usually because it detected they already exist).",
            4: "Accounts where FanDuel warned 'We found another account'.",
            5: "Accounts where FanDuel's servers were down.",
            6: "Accounts where FanDuel couldn't verify the personal data."
        }
        
        self.notebook_tooltip = ToolTip(self.notebook, "")
        self.current_tab_index = -1
        def on_notebook_motion(event):
            try:
                index = self.notebook.index(f"@{event.x},{event.y}")
                if index in tab_tooltips:
                    if self.current_tab_index != index:
                        self.current_tab_index = index
                        self.notebook_tooltip.text = tab_tooltips[index]
                        self.notebook_tooltip.enter()
                else:
                    self.current_tab_index = -1
                    self.notebook_tooltip.leave()
            except tk.TclError:
                self.current_tab_index = -1
                self.notebook_tooltip.leave()
                
        self.notebook.bind("<Motion>", on_notebook_motion)
        self.notebook.bind("<Leave>", lambda e: self.notebook_tooltip.leave())
        self.notebook.bind("<ButtonPress>", lambda e: self.notebook_tooltip.leave())
        
        # --- TREEVIEWS ---
        columns_pending = ("ID", "Email", "Username", "Status")
        columns_done = ("ID", "Email", "Username", "Time Ran", "Status")
        
        self.tree_pending = self.create_treeview(self.tab_pending, columns_pending, "pending")
        self.tree_created = self.create_treeview(self.tab_created, columns_done, "created")
        self.tree_failed = self.create_treeview(self.tab_failed, columns_done + ("Reason",), "failed")
        self.tree_skipped = self.create_treeview(self.tab_skipped, columns_done + ("Reason",), "skipped")
        self.tree_another_account = self.create_treeview(self.tab_another_account, columns_done + ("Reason",), "another_account")
        self.tree_service_unavailable = self.create_treeview(self.tab_service_unavailable, columns_done + ("Reason",), "service_unavailable")
        self.tree_unable_to_verify = self.create_treeview(self.tab_unable_to_verify, columns_done + ("Reason",), "unable_to_verify")
        
        self.tabs_info = [
            ("pending", self.tab_pending, self.tree_pending, columns_pending),
            ("created", self.tab_created, self.tree_created, columns_done),
            ("failed", self.tab_failed, self.tree_failed, columns_done + ("Reason",)),
            ("skipped", self.tab_skipped, self.tree_skipped, columns_done + ("Reason",)),
            ("another_account", self.tab_another_account, self.tree_another_account, columns_done + ("Reason",)),
            ("service_unavailable", self.tab_service_unavailable, self.tree_service_unavailable, columns_done + ("Reason",)),
            ("unable_to_verify", self.tab_unable_to_verify, self.tree_unable_to_verify, columns_done + ("Reason",)),
        ]

    def create_treeview(self, parent, columns, tab_name):
        # --- Frame for controls and treeview ---
        container = tk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True)
        
        # --- Control Bar --- 
        control_bar = tk.Frame(container, bg="#ffffff", pady=5, padx=10)
        control_bar.pack(side=tk.TOP, fill=tk.X)
        
        lbl = tk.Label(control_bar, text="Filter by Date:", font=("Segoe UI", 10, "bold"), bg="#ffffff")
        lbl.pack(side=tk.LEFT, padx=5)
        add_tooltip(lbl, "Select a date to view accounts that were run on that specific day")
        
        date_var = tk.StringVar(value="All Dates")
        date_combo = ttk.Combobox(control_bar, textvariable=date_var, state="readonly", width=20)
        date_combo.pack(side=tk.LEFT, padx=5)
        date_combo.bind("<<ComboboxSelected>>", lambda e, t=tab_name, v=date_var: self.filter_by_date(t, v))
        add_tooltip(date_combo, "Select a date to view accounts that were run on that specific day")
        
        btn_json = tk.Button(control_bar, text="📥 Export JSON", command=lambda t=tab_name: self.export_data(t, "json"), 
                  bg="#e2e8f0", font=("Segoe UI", 9, "bold"), relief=tk.FLAT, padx=10)
        btn_json.pack(side=tk.RIGHT, padx=5)
        add_tooltip(btn_json, "Download the current list of accounts as a JSON file")
        
        btn_csv = tk.Button(control_bar, text="📥 Export CSV", command=lambda t=tab_name: self.export_data(t, "csv"), 
                  bg="#e2e8f0", font=("Segoe UI", 9, "bold"), relief=tk.FLAT, padx=10)
        btn_csv.pack(side=tk.RIGHT, padx=5)
        add_tooltip(btn_csv, "Download the current list of accounts as a CSV file")
        
        btn_excel = tk.Button(control_bar, text="📥 Export Excel", command=lambda t=tab_name: self.export_data(t, "excel"), 
                  bg="#e2e8f0", font=("Segoe UI", 9, "bold"), relief=tk.FLAT, padx=10)
        btn_excel.pack(side=tk.RIGHT, padx=5)
        add_tooltip(btn_excel, "Download the current list of accounts as an Excel (.xlsx) file")
        
        if tab_name not in ["pending", "created"]:
            btn_mark = tk.Button(control_bar, text="Mark All & Send to Pending", command=lambda t=tab_name: self.send_to_pending(t), 
                      bg="#ffc107", font=("Segoe UI", 9, "bold"), relief=tk.FLAT, padx=10)
            btn_mark.pack(side=tk.RIGHT, padx=5)
            add_tooltip(btn_mark, "Move all accounts in this tab back to the Pending tab so the bot can try them again")
        
        # --- Treeview --- 
        tree = ttk.Treeview(container, columns=columns, show="headings", selectmode="extended")
        for col in columns:
            tree.heading(col, text=col)
            if col == "ID":
                tree.column(col, width=50, anchor=tk.CENTER)
            elif col == "Status":
                tree.column(col, width=100, anchor=tk.CENTER)
            elif col == "Time Ran":
                tree.column(col, width=150, anchor=tk.CENTER)
            elif col == "Reason":
                tree.column(col, width=300)
            else:
                tree.column(col, width=200)
            
        scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        
        # Alternate row colors
        tree.tag_configure('evenrow', background='#f9fafb')
        tree.tag_configure('oddrow', background='#ffffff')
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Context menu (Right-click)
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="✏ Edit Account", command=lambda: self.edit_account(tree))
        menu.add_command(label="🗑 Delete Account", command=lambda: self.delete_account(tree))
        menu.add_separator()
        menu.add_command(label="▶ Run This Account", command=lambda: self.run_single(tree))
        menu.add_command(label="🖼 View Screenshot", command=lambda: self.view_screenshot(tree))
        if tab_name not in ["pending", "created"]:
            menu.add_separator()
            menu.add_command(label="Send to Pending", command=lambda t=tab_name, tr=tree: self.send_selected_to_pending(t, tr))
        
        def show_menu(event):
            item = tree.identify_row(event.y)
            if item:
                if item not in tree.selection():
                    tree.selection_set(item)
                menu.post(event.x_root, event.y_root)
                
        tree.bind("<Button-3>", show_menu)
        tree.bind("<Double-1>", lambda e: self.view_screenshot(tree))
        
        # Store date combo for this tab
        self.__dict__[f"date_combo_{tab_name}"] = date_combo
        self.__dict__[f"date_var_{tab_name}"] = date_var
        
        return tree

    def refresh_lists(self):
        # Clear all trees
        trees = [self.tree_pending, self.tree_created, self.tree_failed, self.tree_skipped, 
                 self.tree_another_account, self.tree_service_unavailable, self.tree_unable_to_verify]
        for tree in trees:
            for item in tree.get_children():
                tree.delete(item)
                
        # Collect all unique dates
        dates = set()
        for idx, acc in enumerate(self.accounts):
            timestamp = acc.get("timestamp", "-")
            if timestamp != "-":
                date_part = timestamp.split(" ")[0]
                dates.add(date_part)
        
        sorted_dates = sorted(dates, reverse=True)
        
        # Update all date combos
        for tab_name, _, _, _ in self.tabs_info:
            combo = self.__dict__[f"date_combo_{tab_name}"]
            current_val = self.__dict__[f"date_var_{tab_name}"].get()
            combo['values'] = ["All Dates"] + sorted_dates
            if current_val not in combo['values']:
                self.__dict__[f"date_var_{tab_name}"].set("All Dates")
        
        # Now populate the trees
        for idx, acc in enumerate(self.accounts):
            status = classify_account(acc)
            timestamp = acc.get("timestamp", "-")
                
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            if status == "Pending":
                self.tree_pending.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), status], tags=(tag,))
            elif status == "Created":
                self.tree_created.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status], tags=(tag,))
            elif status == "Failed":
                self.tree_failed.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status, acc.get("reason", "")], tags=(tag,))
            elif status == "Skipped":
                self.tree_skipped.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status, acc.get("reason", "")], tags=(tag,))
            elif status == "Another Account":
                self.tree_another_account.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status, acc.get("reason", "")], tags=(tag,))
            elif status == "Service Unavailable":
                self.tree_service_unavailable.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status, acc.get("reason", "")], tags=(tag,))
            elif status == "Unable to Verify":
                self.tree_unable_to_verify.insert("", tk.END, values=[idx, acc.get("email", ""), acc.get("username", ""), timestamp, status, acc.get("reason", "")], tags=(tag,))
    
    def get_accounts_for_tab(self, tab_name):
        accs = []
        for idx, acc in enumerate(self.accounts):
            status = classify_account(acc)
            
            if status == TAB_STATUS_MAP[tab_name]:
                acc_copy = acc.copy()
                acc_copy["id"] = idx
                accs.append(acc_copy)
        return accs
    
    def filter_by_date(self, tab_name, date_var):
        # First refresh all trees, then filter the selected tab's tree
        self.refresh_lists()
        
        filter_date = date_var.get()
        if filter_date == "All Dates":
            return
            
        # Find the tree for this tab
        tree = None
        for name, _, tr, _ in self.tabs_info:
            if name == tab_name:
                tree = tr
                break
        
        if not tree:
            return
            
        # Filter the tree
        for item in tree.get_children():
            values = tree.item(item, "values")
            if len(values) > 3: # has Time Ran
                timestamp = values[3]
                if timestamp.startswith(filter_date):
                    continue
            tree.delete(item)
    
    def export_data(self, tab_name, export_format):
        accounts = self.get_accounts_for_tab(tab_name)
        
        # Filter by date
        filter_date = self.__dict__[f"date_var_{tab_name}"].get()
        if filter_date != "All Dates":
            filtered = []
            for acc in accounts:
                timestamp = acc.get("timestamp", "-")
                if timestamp.startswith(filter_date):
                    filtered.append(acc)
            accounts = filtered
        
        if not accounts:
            messagebox.showinfo("Notice", "No data to export.")
            return
        
        filetypes = {
            "json": [("JSON Files", "*.json")],
            "csv": [("CSV Files", "*.csv")],
            "excel": [("Excel Files", "*.xlsx")]
        }
        
        ext = "json" if export_format == "json" else "csv" if export_format == "csv" else "xlsx"
        file_path = filedialog.asksaveasfilename(defaultextension=f".{ext}", filetypes=filetypes[export_format])
        if not file_path:
            return
        
        try:
            if export_format == "json":
                with open(file_path, "w") as f:
                    json.dump(accounts, f, indent=4)
            elif export_format == "csv":
                # Get all possible keys from accounts
                keys = set()
                for acc in accounts:
                    keys.update(acc.keys())
                keys = sorted(list(keys))
                
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=keys)
                    writer.writeheader()
                    for acc in accounts:
                        writer.writerow(acc)
            elif export_format == "excel":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = tab_name.capitalize()
                
                # Get all possible keys from accounts
                keys = set()
                for acc in accounts:
                    keys.update(acc.keys())
                keys = sorted(list(keys))
                
                # Write headers
                for col, key in enumerate(keys, 1):
                    cell = ws.cell(row=1, column=col, value=key)
                    cell.font = Font(bold=True)
                
                # Write data
                for row, acc in enumerate(accounts, 2):
                    for col, key in enumerate(keys, 1):
                        ws.cell(row=row, column=col, value=acc.get(key, ""))
                
                wb.save(file_path)
            
            messagebox.showinfo("Done", f"Data exported successfully to {file_path}")
        except Exception as e:
            logger.exception("Failed to export %s data", tab_name)
            messagebox.showerror("Action needed", f"Failed to export data: {e}")
    
    def send_to_pending(self, tab_name):
        if not messagebox.askyesno("Confirm action", "Mark all accounts in this tab as pending and reset?"):
            return
        
        accounts = self.get_accounts_for_tab(tab_name)
        for acc_copy in accounts:
            idx = acc_copy["id"]
            reset_to_pending(self.accounts[idx])
        
        self.save_accounts()
        self.refresh_lists()
        messagebox.showinfo("Done", "All accounts sent to Pending.")
    
    def send_selected_to_pending(self, tab_name, tree):
        indices = self.get_selected_indices(tree)
        if not indices:
            return
        
        if not messagebox.askyesno("Confirm action", f"Mark {len(indices)} account(s) as pending and reset?"):
            return
        
        for idx in indices:
            reset_to_pending(self.accounts[idx])
        
        self.save_accounts()
        self.refresh_lists()
        messagebox.showinfo("Done", "Selected accounts sent to Pending.")

    def open_settings(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Configuration Settings")
        dialog.geometry("800x600")
        
        settings = load_settings()
        
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        tab_general = ttk.Frame(notebook)
        tab_referrals = ttk.Frame(notebook)
        tab_proxies = ttk.Frame(notebook)
        
        notebook.add(tab_general, text="General")
        notebook.add(tab_referrals, text="Referrals")
        notebook.add(tab_proxies, text="Proxies")
        
        # General Tab
        tk.Label(tab_general, text="Referral Mode:", font=("Segoe UI", 10, "bold")).pack(pady=(20, 5))
        mode_var = tk.StringVar(value=settings.get("referral_mode", "rotate"))
        modes = [
            ("Rotate (A -> B -> C -> A)", "rotate", "Runs accounts sequentially through all enabled referrals one by one."),
            ("Sequential 60 mins (One for 60m, then next)", "sequential_60m", "Sticks to one referral for 60 minutes, then switches to the next available one."),
            ("Random Mix (Random but equal distribution)", "random_mix", "Randomly selects a referral for each account, ensuring equal usage over 24 hours."),
            ("Percentage Allocation (Based on weights)", "percentage_allocation", "Allocates accounts to referrals based on the 'percentage' value set in the Referrals tab.")
        ]
        for text, val, desc in modes:
            tk.Radiobutton(tab_general, text=text, variable=mode_var, value=val).pack(anchor=tk.W, padx=20)
            tk.Label(tab_general, text=desc, font=("Segoe UI", 8), fg="#666666").pack(anchor=tk.W, padx=40, pady=(0, 10))
            
        tk.Frame(tab_general, height=2, bg="#cbd5e1").pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(tab_general, text="Edge Browser Path:", font=("Segoe UI", 10, "bold")).pack(pady=(10, 5))
        tk.Label(tab_general, text=r"Example: C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe", font=("Segoe UI", 8), fg="#666666").pack(anchor=tk.W, padx=40)
        
        edge_path_var = tk.StringVar(value=settings.get("edge_path", r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"))
        tk.Entry(tab_general, textvariable=edge_path_var, width=70).pack(anchor=tk.W, padx=40, pady=(0, 10))
            
        # Referrals Tab
        ref_text = tk.Text(tab_referrals, height=20)
        ref_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        ref_str = json.dumps(settings.get("referrals", []), indent=4)
        ref_text.insert(tk.END, ref_str)
        tk.Label(tab_referrals, text="Enter Referrals: Wrap in quotes. When appending, ALWAYS add a comma after the previous one:\n\"www.old.com\", \"www.new.com\"").pack(pady=5)
        
        # Proxies Tab
        proxy_text = tk.Text(tab_proxies, height=20)
        proxy_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        proxy_str = json.dumps(settings.get("proxies", []), indent=4)
        proxy_text.insert(tk.END, proxy_str)
        tk.Label(tab_proxies, text="Enter Proxies: Wrap in quotes. When appending, ALWAYS add a comma after the previous one:\n\"old:proxy:1:1\", \"new:proxy:2:2\"").pack(pady=5)
        
        def save_settings_ui():
            ref_raw = ref_text.get("1.0", tk.END).strip()
            proxy_raw = proxy_text.get("1.0", tk.END).strip()

            try:
                referrals = parse_referrals(ref_raw)
            except ValueError as e:
                messagebox.showerror("Action needed", str(e))
                return

            try:
                proxies = parse_proxies(proxy_raw)
            except ValueError as e:
                messagebox.showerror("Action needed", str(e))
                return
                
            settings["referrals"] = referrals
            settings["proxies"] = proxies
            settings["referral_mode"] = mode_var.get()
            settings["edge_path"] = edge_path_var.get().strip()
            
            # Clean up old keys
            if "urls" in settings:
                del settings["urls"]
            
            save_settings(settings)
            messagebox.showinfo("Done", "Settings saved successfully.")
            dialog.destroy()
            
        tk.Button(dialog, text="Save Settings", command=save_settings_ui, bg="#28a745", fg="white", font=("Segoe UI", 10, "bold")).pack(pady=10)

    def upload_json(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
        if file_path:
            try:
                with open(file_path, "r") as f:
                    new_accs = json.load(f)
                    if isinstance(new_accs, dict):
                        new_accs = [new_accs]
                    if isinstance(new_accs, list):
                        self.accounts.extend(new_accs)
                        self.save_accounts()
                        self.refresh_lists()
                        messagebox.showinfo("Done", f"Appended {len(new_accs)} accounts.")
                    else:
                        messagebox.showerror("Action needed", "JSON must contain a list of accounts.")
            except Exception as e:
                logger.exception("Failed to read account JSON")
                messagebox.showerror("Action needed", f"Failed to read file: {e}")

    def upload_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip() if cell.value else "")
                
                if not headers or not any(headers):
                    messagebox.showerror("Action needed", "No headers found in the first row of the Excel file.")
                    return
                
                new_accs = []
                # Map common non-tech user column names to exact JSON keys
                key_map = {
                    "firstname": "firstName", "first": "firstName",
                    "middlename": "middleName", "middle": "middleName",
                    "lastname": "lastName", "last": "lastName",
                    "referralurl": "referral_url", "referral": "referral_url",
                    "zipcode": "postcode", "zip": "postcode", "postalcode": "postcode",
                    "state": "province"
                }
                valid_exact_keys = ["email", "password", "apt", "address", "city", "province", "postcode", "month", "day", "year", "phone", "username"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    acc = {}
                    has_data = False
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i]:
                            raw_key = headers[i]
                            clean_key = raw_key.lower().replace(" ", "").replace("_", "")
                            
                            if clean_key in key_map:
                                final_key = key_map[clean_key]
                            elif clean_key in valid_exact_keys:
                                final_key = clean_key
                            else:
                                final_key = raw_key # Fallback
                                
                            val_str = str(val).strip() if val is not None else ""
                            if val_str:
                                has_data = True
                            acc[final_key] = val_str
                    
                    if has_data:
                        # Clean up numbers parsed as floats from Excel (e.g. '11.0' -> '11')
                        for k in ["month", "day", "year", "phone", "postcode", "apt"]:
                            if k in acc and acc[k].endswith(".0"):
                                acc[k] = acc[k][:-2]
                        new_accs.append(acc)
                
                if new_accs:
                    self.accounts.extend(new_accs)
                    self.save_accounts()
                    self.refresh_lists()
                    messagebox.showinfo("Done", f"Appended {len(new_accs)} accounts from Excel.")
                else:
                    messagebox.showinfo("Notice", "No valid data found in the Excel file.")
            except Exception as e:
                logger.exception("Failed to read account Excel file")
                messagebox.showerror("Action needed", f"Failed to read Excel file: {e}")

    def get_selected_indices(self, tree):
        selected = tree.selection()
        return [int(tree.item(item, "values")[0]) for item in selected]

    def delete_account(self, tree):
        indices = self.get_selected_indices(tree)
        if not indices: return
        if messagebox.askyesno("Confirm action", f"Delete {len(indices)} account(s)?"):
            # Delete in reverse order to keep indices intact during deletion
            for idx in sorted(indices, reverse=True):
                del self.accounts[idx]
            self.save_accounts()
            self.refresh_lists()

    def add_account(self):
        self.open_editor_dialog(None)

    def edit_account(self, tree):
        indices = self.get_selected_indices(tree)
        if indices:
            self.open_editor_dialog(indices[0])

    def open_editor_dialog(self, idx):
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Account" if idx is not None else "Add Account")
        dialog.geometry("650x500")
        
        fields = [
            "email", "username", "password", "firstName", "middleName", "lastName", 
            "apt", "address", "city", "province", "postcode", "phone", "month", "day", "year", "referral_url"
        ]
        entries = {}
        acc_data = self.accounts[idx] if idx is not None else automate2.DEFAULT_CONFIG.copy()
        
        # Two-column layout
        for i, field in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            tk.Label(dialog, text=field.capitalize() + ":").grid(row=row, column=col, padx=10, pady=10, sticky=tk.E)
            ent = tk.Entry(dialog, width=25)
            ent.insert(0, str(acc_data.get(field, "")))
            ent.grid(row=row, column=col+1, padx=10, pady=10)
            entries[field] = ent
            
        def save(make_pending=False):
            for field, ent in entries.items():
                acc_data[field] = ent.get()
            
            if make_pending:
                acc_data["ran"] = False
                acc_data["success"] = False
                acc_data["reason"] = ""
                
            if idx is None:
                self.accounts.append(acc_data)
            else:
                self.accounts[idx] = acc_data
                
            self.save_accounts()
            self.refresh_lists()
            dialog.destroy()
            
        btn_frame = tk.Frame(dialog)
        btn_frame.grid(row=len(fields)//2 + 1, column=0, columnspan=4, pady=20)
        
        tk.Button(btn_frame, text="Save", command=lambda: save(False), width=15).pack(side=tk.LEFT, padx=10)
        if idx is not None:
            tk.Button(btn_frame, text="Save & Reset to Pending", command=lambda: save(True), width=20).pack(side=tk.LEFT, padx=10)

    def view_screenshot(self, tree):
        indices = self.get_selected_indices(tree)
        if not indices: return
        acc = self.accounts[indices[0]]
        path = acc.get("screenshot")
        screenshot_path = workspace_path(path) if path else None
        
        if screenshot_path and screenshot_path.exists():
            top = tk.Toplevel(self.root)
            top.title(f"Screenshot - {acc.get('email')}")
            try:
                img = Image.open(screenshot_path)
                # Resize to fit screen reasonably
                img.thumbnail((900, 700))
                photo = ImageTk.PhotoImage(img)
                lbl = tk.Label(top, image=photo)
                lbl.image = photo
                lbl.pack()
            except Exception as e:
                messagebox.showerror("Action needed", f"Cannot load image: {e}")
        else:
            messagebox.showinfo("Screenshot not found", "No screenshot available for this account.")

    def stop_automation(self):
        if self.is_running:
            self.stop_requested = True
            # Set the stop flag in automate2
            automate2.set_stop_requested(True)
            self.status_lbl.config(text="Stopping... Closing browser...", fg="#dc3545")
            
            # Kill the browser using automate2's helper function
            automate2.kill_browser()
            
            # Trigger PyAutoGUI's Failsafe to interrupt the running script immediately
            try:
                import pyautogui
                pyautogui.FAILSAFE = True
                pyautogui.moveTo(0, 0)
            except Exception:
                pass

    def run_pending(self):
        indices = [i for i, acc in enumerate(self.accounts) if not acc.get("ran")]
        self.start_runner(indices)

    def run_selected(self):
        # Only run selected items from the pending tab
        indices = self.get_selected_indices(self.tree_pending)
        if indices:
            self.start_runner(indices)
        else:
            messagebox.showinfo("Notice", "Select accounts from the Pending tab (Ctrl+Click to select multiple).")

    def run_single(self, tree):
        indices = self.get_selected_indices(tree)
        if indices:
            self.start_runner([indices[0]])

    def start_runner(self, indices):
        if self.is_running:
            messagebox.showwarning("Please wait", "Automation is already running!")
            return
        if not indices:
            messagebox.showinfo("Notice", "No accounts to run.")
            return
            
        self.is_running = True
        self.stop_requested = False
        # Reset the stop flag in automate2
        automate2.set_stop_requested(False)
        self.stop_btn.config(state=tk.NORMAL, bg="#dc3545", cursor="hand2")
        self.warning_lbl.config(text="Do not touch mouse or keyboard while automation is running.")
        
        # Run in background thread to keep UI responsive
        self.thread = threading.Thread(target=self.runner_thread, args=(indices,), daemon=True)
        logger.info("Starting automation runner for %s account(s)", len(indices))
        self.thread.start()

    def runner_thread(self, indices):
        for i, idx in enumerate(indices):
            if self.stop_requested:
                break
                
            acc = self.accounts[idx]
            email = acc.get('email', 'Unknown')
            
            self.root.after(0, lambda e=email, c=i+1, t=len(indices): 
                            self.status_lbl.config(text=f"Validating: {e} ({c}/{t})", fg="#0056b3"))
            
            current_config = prepare_account_config(acc, automate2.DEFAULT_CONFIG)
            
            is_valid, val_reason = automate2.validate_account(current_config)
            if not is_valid:
                acc["ran"] = True
                acc["success"] = False
                acc["skipped"] = True
                acc["reason"] = val_reason
                acc["timestamp"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.save_accounts()
                self.root.after(0, self.refresh_lists)
                continue
                
            settings = load_settings()
            url, proxy, wait_time = automate2.get_next_url_and_proxy(settings, current_config)
            
            if wait_time > 0:
                for remaining in range(int(wait_time), 0, -1):
                    if self.stop_requested:
                        break
                    mins, secs = divmod(remaining, 60)
                    time_str = f"{mins:02d}:{secs:02d}"
                    self.root.after(0, lambda t=time_str: self.status_lbl.config(text=f"Proxy cooldown: {t}", fg="#f39c12"))
                    time.sleep(1)
                    
            if self.stop_requested:
                break
                
            self.root.after(0, lambda e=email, c=i+1, t=len(indices): 
                            self.status_lbl.config(text=f"Running: {e} ({c}/{t})", fg="#0056b3"))
            
            try:
                is_created, status = automate2.main(current_config, url, proxy)
                outcome = outcome_from_result(is_created, status)
            except Exception as e:
                if self.stop_requested:
                    self.root.after(0, lambda: self.status_lbl.config(text="Stopped by user", fg="#dc3545"))
                    break
                    
                outcome = outcome_from_exception(e)
                logger.exception("Automation failed for %s", email)
                
            screenshot_path = automate2.take_result_screenshot(outcome.screenshot_prefix)
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            apply_outcome_to_account(
                acc,
                outcome,
                screenshot_path,
                username=current_config["username"],
                timestamp=timestamp,
            )
            
            self.save_accounts()
            logger.info("Account finished: email=%s success=%s created=%s reason=%s", email, outcome.success, outcome.created, outcome.reason)
            self.root.after(0, self.refresh_lists)
            
            # Close browser
            automate2.kill_browser()
            
            # Wait period between accounts
            if i < len(indices) - 1 and not self.stop_requested:
                settings = load_settings()
                proxies = settings.get("proxies", [])
                if len(proxies) <= 1:
                    wait_time_minutes = 10 if outcome.success else 11
                    total_seconds = wait_time_minutes * 60
                else:
                    total_seconds = 120 # Minimum 2 minutes buffer when multiple proxies exist
                
                # Custom wait loop with countdown display
                for remaining in range(total_seconds, 0, -1):
                    if self.stop_requested:
                        break
                    
                    mins, secs = divmod(remaining, 60)
                    time_str = f"{mins:02d}:{secs:02d}"
                    self.root.after(0, lambda t=time_str: self.status_lbl.config(text=f"Waiting for next account: {t}", fg="#f39c12"))
                    time.sleep(1)

        # Cleanup after loop finishes or stops
        self.is_running = False
        self.stop_requested = False
        self.root.after(0, lambda: self.stop_btn.config(state=tk.DISABLED, bg="#e2e8f0", cursor="arrow"))
        self.root.after(0, lambda: self.warning_lbl.config(text=""))
        self.root.after(0, lambda: self.status_lbl.config(text="Idle", fg="#28a745"))
        logger.info("Automation queue finished")
        self.root.after(0, lambda: messagebox.showinfo("Done", "Automation queue finished."))

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoUI(root)
    root.mainloop()
